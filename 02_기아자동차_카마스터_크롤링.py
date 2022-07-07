import time
import pyautogui
import requests as rq
from bs4 import BeautifulSoup as bs
import json
import re
from openpyxl import Workbook
import os
import pymysql
from typing import  Union , Optional
import aiohttp
import asyncio

class Application:
    def __init__(self):
        # headers
        self.headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.57 Whale/3.14.133.23 Safari/537.36'}

        # 카운트
        self.employees_count : int = 1

        # 페이지 카운트
        self.page_cnt : int = 1

        # BASE URL
        self.urls = [
            f"https://www.kia.com/api/kia_korea/base/br02/branchEmployee.selectNewBranchEmployeeList?pageNum={page}&sc.searchKey%5B2%5D=&sc.searchType%5B2%5D=employeeName&sc.tel=&sortKey%5B0%5D=employeeNm&sortKey%5B1%5D=&sortType%5B0%5D=A&sortType%5B1%5D="
            for page in range(1, self.page_cnt + 1)]

    def run(self) -> list:
        # Session 객체 생성
        with rq.Session() as session :
            # List Comprehension 문법 적용
            result = [self.fetch(session=session,url=url)for url in self.urls]

        return result

    def fetch(self,session,url: str)-> list:
        # with Context Manager로 url 세션 Open
        with session.get(url,headers=self.headers) as response :
            result = response.text

            # json 데이터로 형변환
            json_data = json.loads(result)

            # Key : dataInfo의 Value값 추출
            data_info = json_data['dataInfo']

            # 추출데이터 담을 리스트 변수 선언
            info_list : list = []

            for data in data_info:
                employeeNm : str = data['employeeNm']  # 이름
                employeeTel : str  = data['tel']  # 핸드폰번호
                employeeEmail : str = data['email']  # 이메일주소
                greeting : str = data['greeting']  # 자기소개말
                branchNm : str = data['branchNm']  # 지역이름
                typeName : str = data['typeName']  # 지점인지 대리점인지
                branch : str = branchNm + typeName # 지점명

                # 전화번호 없는 경우 빈 공란 처리
                if employeeTel == None or employeeTel == "":
                    employeeTel = ""

                # 이메일 없는 경우 빈 공란 처리
                if employeeEmail == None or employeeEmail == "":
                    employeeEmail = ""

                # 소개말 없는 경우 빈 공란 처리
                if greeting == None or greeting == "":
                    greeting = ""
                else:
                    greeting = re.sub('[\r\n\t]', '', greeting)
                    greeting = greeting.strip()
                    greeting = re.sub('[^가-힣ㄱ-ㅎA-Za-z0-9]', '', greeting)  # DB 저장위해 데이터 전처리하였음.

                info_list.append([self.employees_count, employeeNm, employeeTel, employeeEmail, branch, greeting])
                print(f"사원번호 : {self.employees_count}\n사원이름 : {employeeNm}\n전화번호 : {employeeTel}\n이메일 : {employeeEmail}\n지점 : {branch}\n소개말 : {greeting}\n")

                # 사원번호 증감시키기
                self.employees_count += 1

            return info_list

class OpenPyXL:
    def __init__(self):
        self.results : list = Application().run()

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = '기아자동차_영업사원'
        self.ws.append(['사원번호','사원이름','전화번호','이메일','지점','소개말'])

        self.ws.column_dimensions['A'].width = 10
        self.ws.column_dimensions['B'].width = 15
        self.ws.column_dimensions['C'].width = 40
        self.ws.column_dimensions['D'].width = 60
        self.ws.column_dimensions['E'].width = 60
        self.ws.column_dimensions['F'].width = 90

    def savefile(self) -> None:
        row : int = 2
        for x in self.results :
            for result in x :
                print(result)

            self.ws[f"A{row}"] = result[0]
            self.ws[f"B{row}"] = result[1]
            self.ws[f"C{row}"] = result[2]
            self.ws[f"D{row}"] = result[3]
            self.ws[f"E{row}"] = result[4]
            self.ws[f"F{row}"] = result[-1]

            row += 1

        savePath : str = os.path.abspath('기아자동차_영업사원')
        fileName : str = '영업사원정보.xlsx'
        if not os.path.exists(savePath) :
            os.mkdir(savePath)

        self.wb.save(os.path.join(savePath,fileName))
        self.wb.close()

        pyautogui.alert(f'파일 저장 완료\n\n{savePath}')

if __name__ == "__main__" :
    app = OpenPyXL()

    app.savefile()